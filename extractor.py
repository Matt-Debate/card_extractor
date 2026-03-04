import io
import csv
import os
import re
from urllib.parse import urlparse
from docx.enum.text import WD_COLOR_INDEX
from openpyxl import Workbook
from openpyxl.styles import Alignment


def is_heading(para, level_or_levels):
    name = (para.style.name or "").lower()
    if isinstance(level_or_levels, (list, tuple, set)):
        levels = level_or_levels
    else:
        levels = [level_or_levels]
    for level in levels:
        if f"heading {level}" in name or name == f"heading{level}":
            return True
    return False

def normalize_separators(s: str) -> str:
    dash_chars = "\u2010\u2011\u2012\u2013\u2014\u2015\u2212"
    s = re.sub(f"[{re.escape(dash_chars)}]", "-", s)
    s = s.replace("~~", "-")
    s = re.sub(r"-{2,}", "-", s)
    s = s.strip(" -")
    return s


def strip_numeric_prefix_tokens(tokens):
    out = tokens[:]
    while out and re.fullmatch(r"\[?\d+\]?", out[0]):
        out.pop(0)
    return out


def tournament_short_name(raw: str) -> str:
    r = raw.strip()
    r_low = r.lower()
    if "university-of-pennsylvania" in r_low or re.search(r"\bupenn\b", r_low):
        if "liberty-bell" in r_low:
            return "UPenn Liberty Bell"
        return "UPenn"
    if "stanford" in r_low:
        if "invitational" in r_low or "invitationals" in r_low:
            return "Stanford Invitational"
        if "online" in r_low:
            return "Stanford Online Invitational"
        return "Stanford"
    if "bellaire" in r_low:
        return "Bellaire"
    if "pennsbury" in r_low:
        return "Pennsbury"
    if "unlv" in r_low or "golden-desert" in r_low:
        return "UNLV Golden Desert"
    cleaned = re.sub(r"\s+", " ", r.replace("-", " ")).strip()
    return " ".join(w.upper() if w.isupper() and len(w) <= 5 else w.capitalize() for w in cleaned.split())


ROUND_NUM_RE = re.compile(r"(?:^|-)Round-(\d+)$", re.IGNORECASE)
SPECIAL_ROUNDS = {
    "Octas",
    "Quarters",
    "Semis",
    "Finals",
    "Doubles",
    "Triples",
    "All-Rounds",
    "All",
}


def split_school_team_side(name_no_ext):
    s = normalize_separators(name_no_ext)
    parts = [p for p in s.split("-") if p]
    if len(parts) < 4:
        return None, None, None, ""
    school, team_code, side = parts[0], parts[1], parts[2]
    rest = "-".join(parts[3:])
    return school, team_code, side, rest


def parse_round_and_tournament(rest):
    rest_norm = normalize_separators(rest)
    tokens = [t for t in rest_norm.split("-") if t]
    tokens = strip_numeric_prefix_tokens(tokens)
    if not tokens:
        return None, None, None

    m = ROUND_NUM_RE.search("-" + tokens[-2] + "-" + tokens[-1]) if len(tokens) >= 2 else None
    if m:
        round_number = int(m.group(1))
        tournament_tokens = tokens[:-2]
        tournament_raw = "-".join(tournament_tokens) if tournament_tokens else None
        return tournament_raw, round_number, "Round"

    last = tokens[-1]
    if last in SPECIAL_ROUNDS:
        tournament_tokens = tokens[:-1]
        tournament_raw = "-".join(tournament_tokens) if tournament_tokens else None
        return tournament_raw, None, last

    return "-".join(tokens), None, None


def parse_filename_metadata(source_name):
    if ":" in source_name:
        source_name = source_name.split(":", 1)[1]
    base = os.path.basename(source_name)
    name_no_ext = os.path.splitext(base)[0]

    school, team_code, side, rest = split_school_team_side(name_no_ext)
    side_norm = None
    if side:
        side_upper = side.strip().lower()
        if side_upper in ("pro", "con"):
            side_norm = side_upper.capitalize()
        else:
            side_norm = side

    tournament_raw, round_number, round_label = parse_round_and_tournament(rest)
    tournament = tournament_short_name(tournament_raw) if tournament_raw else None

    if round_number is not None:
        round_value = f"Round {round_number}"
    elif round_label:
        round_value = round_label
    else:
        round_value = None

    return {
        "filename": base,
        "school": school,
        "team_code": team_code,
        "side": side_norm,
        "tournament_raw": tournament_raw,
        "tournament": tournament,
        "round": round_value,
    }


CITATION_MIN_SCORE = 4
ORAL_CITATION_START_RE = re.compile(
    r"^(?:[A-Z][A-Za-z'`.-]{2,}|[A-Z]{2,})(?:\s+(?:[A-Z][A-Za-z'`.-]{1,}|[A-Z]{2,}|&|and)){0,3}\s*(?:\(|,)?\s*(?:'?\d{2}|(?:19|20)\d{2})\b"
)
YEAR_RE = re.compile(r"\b(?:19|20)\d{2}\b")
SHORT_YEAR_RE = re.compile(r"\b'?\d{2}\b")
URL_RE = re.compile(r"https?://|www\.", re.IGNORECASE)
URL_CANDIDATE_RE = re.compile(r"(?:https?://|www\.)[^\s<>\"']+", re.IGNORECASE)
BRACKETED_CITATION_RE = re.compile(r"\[[^\]]{12,}\]")
PAREN_SOURCE_RE = re.compile(r"\([^)]{20,}\)")
CITATION_KEYWORDS_RE = re.compile(
    r"\b(doi|journal|university|press|vol\.?|no\.?|pp\.?|retrieved|accessed|published|updated)\b",
    re.IGNORECASE,
)
NEWSWIRE_DATELINE_RE = re.compile(
    r"^(?:[A-Z][A-Z.'\-\s]{2,},\s*)?(?:(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+\d{1,2},?\s+(?:19|20)\d{2}|[A-Z][a-z]+,\s*(?:19|20)\d{2}|\([A-Z]{2,}\)\s*[—-])"
)
AUTHOR_YEAR_LEAD_ONLY_RE = re.compile(
    r"^(?:\[[A-Z]{1,8}\]\s*)?(?:[A-Z][A-Za-z'`.-]{2,}|[A-Z]{2,})(?:\s+(?:[A-Z][A-Za-z'`.-]{1,}|[A-Z]{2,}|&|and)){0,3}\s*(?:\(|,)?\s*(?:'?\d{2}|(?:19|20)\d{2})\b"
)
URL_START_RE = re.compile(r"^(?:\[)?(?:https?://|www\.)", re.IGNORECASE)
BODY_SPILLOVER_RE = re.compile(
    r"\b(?:however|as a result|therefore|meanwhile|nevertheless|the aff|the neg|affirming|negating|this proves|this means|thus[, ]+the)\b",
    re.IGNORECASE,
)
ROUND_METADATA_RE = re.compile(
    r"\b(?:round\s*\d+|speech and debate|invitational|tournament|aff|neg|vs\.?|1ac|2ac|1nc|2nc|1nr|2nr|1ar|2ar|cross[-\s]?ex)\b",
    re.IGNORECASE,
)
CITATION_AUTHOR_START_RE = re.compile(
    r"\b[A-Z][A-Za-z'`.-]{2,}(?:\s+[A-Z][A-Za-z'`.-]{1,}){0,3}\s*(?:,|\(|\[)?\s*(?:\d{1,2}[/-]\d{1,2}[/-](?:19|20)?\d{2}|(?:19|20)\d{2}|'?\d{2})\b"
)


def _normalize_text(input_text):
    value = input_text or ""
    value = value.replace("\u2018", "'").replace("\u2019", "'")
    value = value.replace("\u201c", '"').replace("\u201d", '"')
    value = re.sub(r"[\u2013\u2014\u2212]", "-", value)
    value = value.replace("\u00a0", " ")
    value = re.sub(r"[\u2000-\u200b]", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def _word_count(text):
    return len([part for part in text.split() if part])


def _first_words(text, count):
    return " ".join(text.split()[:count])


def _looks_like_round_metadata_label(text):
    normalized = _normalize_text(text)
    if not normalized:
        return False
    return bool(ROUND_METADATA_RE.search(normalized))


def _normalize_tag_title_text(text):
    normalized = _normalize_text(text)
    if not normalized:
        return ""
    normalized = re.sub(r"^\s*[-*]+\s*", "", normalized)
    normalized = re.sub(r"https?://\S+|www\.\S+", " ", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized.rstrip(".,;:- ").strip()


def _parse_heading_level(para):
    style_name = (getattr(getattr(para, "style", None), "name", "") or "").strip().lower()
    if not style_name:
        return None
    heading_match = re.search(r"heading\s*([1-9])", style_name)
    if heading_match:
        return int(heading_match.group(1))
    short_match = re.fullmatch(r"h([1-9])", style_name)
    if short_match:
        return int(short_match.group(1))
    return None


def _paragraph_is_bold(para):
    total_chars = 0
    bold_chars = 0
    for run in para.runs:
        run_text = _normalize_text(run.text or "")
        if not run_text:
            continue
        run_len = len(run_text)
        total_chars += run_len
        if bool(run.bold) or bool(run.font.bold):
            bold_chars += run_len
    if total_chars == 0:
        return False
    return (bold_chars / total_chars) >= 0.4


def _is_likely_newswire_dateline(text):
    normalized = _normalize_text(text)
    if not normalized:
        return False
    return bool(NEWSWIRE_DATELINE_RE.match(normalized))


def _parse_inline_tag_citation(text):
    normalized = _normalize_text(text)
    if not normalized:
        return None
    match = re.match(r"^(.{1,160}?)\s*[-–—|:]\s*(.{3,})$", normalized)
    if not match:
        return None
    tag_candidate = _normalize_text(match.group(1))
    citation_tail = _normalize_text(match.group(2))
    if not tag_candidate or not citation_tail:
        return None
    return {"tag_candidate": tag_candidate, "citation_tail": citation_tail}


def _citation_score(text):
    value = _normalize_text(text)
    if not value:
        return 0
    lower = value.lower()
    words = _word_count(value)
    punctuation_count = len(re.findall(r"[.?!]", value))
    has_year = bool(YEAR_RE.search(value))
    has_short_year = bool(SHORT_YEAR_RE.search(value))
    has_url = bool(URL_RE.search(value))
    has_oral_start = bool(ORAL_CITATION_START_RE.search(value))
    has_bracketed_written = bool(BRACKETED_CITATION_RE.search(value))
    has_parenthetical_written = bool(PAREN_SOURCE_RE.search(value))
    has_citation_keywords = bool(CITATION_KEYWORDS_RE.search(value))
    has_comma_year_pattern = bool(re.search(r"[,;:]\s*(?:19|20)\d{2}\b", value))
    is_likely_tag_line = bool(
        re.search(r"^(contention|advantage|framework|overview|case)\b", value, re.IGNORECASE)
    )
    inline_tag_citation = _parse_inline_tag_citation(value)

    score = 0
    if has_year:
        score += 2
    if has_oral_start:
        score += 4
    if has_url:
        score += 3
    if has_bracketed_written:
        score += 4
    if has_parenthetical_written:
        score += 2
    if has_citation_keywords:
        score += 1
    if has_comma_year_pattern:
        score += 1
    if has_oral_start and (has_url or has_bracketed_written or has_parenthetical_written):
        score += 3
    if inline_tag_citation:
        tail = inline_tag_citation["citation_tail"]
        score += 2
        if ORAL_CITATION_START_RE.search(tail):
            score += 3
        if YEAR_RE.search(tail):
            score += 1
        if URL_RE.search(tail) or BRACKETED_CITATION_RE.search(tail) or PAREN_SOURCE_RE.search(tail):
            score += 2
    if words < 2:
        score -= 3
    if words > 70:
        score -= 4
    if len(value) > 600:
        score -= 5
    if not has_oral_start and not has_url and not has_bracketed_written and punctuation_count >= 2:
        score -= 2
    if not has_year and not has_short_year:
        score -= 2
    if has_oral_start and words <= 20 and not has_url and not has_bracketed_written and not has_parenthetical_written:
        score -= 1
    if is_likely_tag_line:
        score -= 2
    if _is_likely_newswire_dateline(value) and not has_url:
        score -= 4
    if lower.startswith("http"):
        score += 1
    return score


def _title_score(text, is_bold):
    value = _normalize_text(text)
    if not value:
        return -3

    score = 0
    words = _word_count(value)
    if is_bold:
        score += 2
    if 3 <= len(value) <= 90:
        score += 1
    if 1 <= words <= 14:
        score += 1
    if not YEAR_RE.search(value):
        score += 1
    else:
        score -= 2
    if len(re.findall(r"[.?!]", value)) <= 1:
        score += 1
    if re.search(r"[.?!]$", value) and not is_bold:
        score -= 2
    if words > 10:
        score -= 1
    if ":" in value and words <= 16:
        score += 1
    if _is_likely_newswire_dateline(value):
        score -= 6
    if URL_RE.search(value):
        score -= 4
    if _citation_score(value) >= CITATION_MIN_SCORE:
        score -= 3
    return score


def _build_paragraph_info(index, para):
    text = _normalize_text(para.text or "")
    heading_level = _parse_heading_level(para)
    is_bold = _paragraph_is_bold(para)
    citation_score = _citation_score(text)
    return {
        "index": index,
        "text": text,
        "is_bold": is_bold,
        "heading_level": heading_level,
        "is_heading_style": heading_level is not None,
        "citation_score": citation_score,
        "title_score": _title_score(text, is_bold),
        "has_url": bool(URL_RE.search(text)),
    }


def _is_likely_citation_paragraph(paragraph, require_url_for_citation):
    if _is_likely_newswire_dateline(paragraph["text"]):
        return False
    if paragraph["citation_score"] < CITATION_MIN_SCORE:
        return False
    if require_url_for_citation and not paragraph["has_url"]:
        return False
    return True


def _split_inline_citation_from_title(text):
    normalized = _normalize_text(text)
    if not normalized:
        return normalized
    inline_split = _parse_inline_tag_citation(normalized)
    if not inline_split:
        return normalized
    left = inline_split["tag_candidate"]
    right = inline_split["citation_tail"]
    if not left or not right:
        return normalized
    if _word_count(left) > 12:
        return normalized
    if _citation_score(right) < CITATION_MIN_SCORE:
        return normalized
    return left


def _select_title_from_pre_citation_range(range_infos, require_url_for_citation):
    if not range_infos:
        return {"title": "", "indexes": []}

    tail = range_infos[max(0, len(range_infos) - 3) :]
    selected = []
    for candidate in reversed(tail):
        candidate_text = candidate["text"]
        if _is_likely_newswire_dateline(candidate_text):
            continue
        if _is_likely_citation_paragraph(candidate, require_url_for_citation):
            break
        if candidate["title_score"] >= 3 or (candidate["is_bold"] and candidate["title_score"] >= 2):
            selected.insert(0, candidate)
            continue
        if (
            selected
            and candidate["title_score"] >= 2
            and len(candidate_text) <= 50
            and not re.search(r"[.?!]$", candidate_text)
        ):
            selected.insert(0, candidate)
            continue
        if selected:
            break

    if not selected:
        fallback = None
        for candidate in reversed(tail):
            if _is_likely_newswire_dateline(candidate["text"]):
                continue
            if _is_likely_citation_paragraph(candidate, require_url_for_citation):
                continue
            if candidate["title_score"] >= 1:
                fallback = candidate
                break
        if not fallback:
            return {"title": "", "indexes": []}
        return {
            "title": _split_inline_citation_from_title(fallback["text"]),
            "indexes": [fallback["index"]],
        }

    title = _normalize_text(
        " ".join(_split_inline_citation_from_title(item["text"]) for item in selected)
    )
    return {"title": title, "indexes": [item["index"] for item in selected]}


def _reconstruct_title_from_body(body_infos):
    first = next(
        (
            item
            for item in body_infos
            if item["text"] and not _is_likely_newswire_dateline(item["text"])
        ),
        None,
    )
    if not first:
        return ""
    return _normalize_text(_first_words(first["text"], 7))


def _looks_like_tag_line(text):
    normalized = _normalize_tag_title_text(text)
    if not normalized:
        return False
    return bool(
        re.match(
            r"^(?:tag|contention|advantage|framework|overview|case|offcase|oncase)\b",
            normalized,
            re.IGNORECASE,
        )
    )


def _find_recent_heading_title(scored, citation_index):
    for item in reversed(scored):
        if item["index"] >= citation_index:
            continue
        if item["heading_level"] not in (1, 2, 3):
            continue
        text = _normalize_tag_title_text(item["text"])
        if not text:
            continue
        if _looks_like_round_metadata_label(text):
            continue
        if _is_likely_citation_paragraph(item, require_url_for_citation=False):
            continue
        return text
    return ""


def _derive_selected_title_and_tag(title_selection, scored_by_index):
    indexes = title_selection.get("indexes", []) or []
    selected_texts = [
        _normalize_tag_title_text(scored_by_index[index]["text"])
        for index in indexes
        if index in scored_by_index and scored_by_index[index]["text"]
    ]
    selected_texts = [text for text in selected_texts if text]
    if not selected_texts:
        return {"title": _normalize_tag_title_text(title_selection.get("title", "")), "tag": ""}

    if len(selected_texts) >= 2:
        maybe_tag = selected_texts[-1]
        maybe_title = _normalize_tag_title_text(" ".join(selected_texts[:-1]))
        if maybe_title and maybe_tag and _looks_like_tag_line(maybe_tag):
            return {"title": maybe_title, "tag": maybe_tag}

    merged = _normalize_tag_title_text(title_selection.get("title", ""))
    if not merged:
        merged = _normalize_tag_title_text(" ".join(selected_texts))
    if _looks_like_tag_line(merged):
        return {"title": "", "tag": merged}
    return {"title": merged, "tag": ""}


def _derive_tag_from_title(title, citation):
    raw = _normalize_tag_title_text(title)
    if not raw:
        return ""
    if _looks_like_round_metadata_label(raw):
        return ""
    words = _word_count(raw)
    if 3 <= words <= 12:
        return raw

    for sep in ("---", "--", " - ", " — ", " – ", ": "):
        idx = raw.find(sep)
        if idx <= 0:
            continue
        left = _normalize_tag_title_text(raw[:idx])
        right = _normalize_tag_title_text(raw[idx + len(sep) :])
        if 3 <= _word_count(left) <= 12 and right and not _looks_like_round_metadata_label(left):
            return left

    citation_head = _normalize_tag_title_text(citation)[:220].lower()
    sentence_like = _normalize_tag_title_text(raw.split(".", 1)[0])
    sentence_words = _word_count(sentence_like)
    if sentence_like and 3 <= sentence_words <= 12 and not citation_head.startswith(sentence_like.lower()):
        return sentence_like

    if words > 12:
        return _first_words(raw, 12)
    return raw


def _derive_from_citation_lead(citation):
    normalized = _normalize_tag_title_text(citation)
    if not normalized:
        return {"title": "", "tag": ""}
    if not URL_RE.search(normalized):
        return {"title": "", "tag": ""}

    author_match = CITATION_AUTHOR_START_RE.search(normalized)
    if not author_match or author_match.start() < 8:
        return {"title": "", "tag": ""}

    lead_raw = _normalize_tag_title_text(
        normalized[: author_match.start()]
        .lstrip("[")
        .lstrip('"\' ')
        .rstrip('"\' ')
    )
    if not lead_raw:
        return {"title": "", "tag": ""}
    lead_words = _word_count(lead_raw)
    if lead_words < 3 or lead_words > 18:
        return {"title": "", "tag": ""}
    if _looks_like_round_metadata_label(lead_raw) or _is_likely_newswire_dateline(lead_raw):
        return {"title": "", "tag": ""}

    tag = _first_words(lead_raw, 12) if lead_words > 12 else lead_raw
    title = _normalize_tag_title_text(lead_raw)
    return {"title": title, "tag": tag}


def _looks_like_author_year_lead(text):
    normalized = _normalize_text(text)
    if not normalized:
        return False
    if URL_RE.search(normalized):
        return False
    if _is_likely_newswire_dateline(normalized):
        return False
    words = _word_count(normalized)
    if words < 2 or words > 16:
        return False
    return bool(AUTHOR_YEAR_LEAD_ONLY_RE.search(normalized))


def _is_url_first_short_citation(text):
    normalized = _normalize_text(text)
    if not normalized:
        return False
    if not URL_START_RE.search(normalized):
        return False
    words = _word_count(normalized)
    return 0 < words <= 18


def _trim_trailing_unmatched_closers(text):
    value = _normalize_text(text)
    left_brackets = value.count("[")
    right_brackets = value.count("]")
    left_parens = value.count("(")
    right_parens = value.count(")")
    while right_brackets > left_brackets and value.endswith("]"):
        value = value[:-1].rstrip()
        right_brackets -= 1
    while right_parens > left_parens and value.endswith(")"):
        value = value[:-1].rstrip()
        right_parens -= 1
    return value


def _repair_missing_trailing_closers(text):
    value = _normalize_text(text)
    left_brackets = value.count("[")
    right_brackets = value.count("]")
    left_parens = value.count("(")
    right_parens = value.count(")")
    if left_brackets - right_brackets == 1 and re.search(r"\[[^\]]{6,}$", value):
        value = f"{value}]"
    if left_parens - right_parens == 1 and re.search(r"\([^)]{6,}$", value):
        value = f"{value})"
    return value


def _strip_double_slash_suffix(text):
    value = _normalize_text(text)
    match = re.search(r"(?:\s|\])\/\/\s*", value)
    if not match:
        return value
    cut_index = match.start()
    if cut_index < 6:
        return value
    trimmed = _normalize_text(value[:cut_index])
    return trimmed if trimmed else value


def _truncate_citation_spillover(text):
    value = _normalize_text(text)
    words = _word_count(value)
    if words < 220:
        return value
    start_index = max(110, int(len(value) * 0.45))
    if start_index >= len(value) - 20:
        return value
    tail = value[start_index:]
    marker_match = BODY_SPILLOVER_RE.search(tail)
    if not marker_match:
        return value
    cut_index = start_index + marker_match.start()
    if cut_index < 120:
        return value
    truncated = _normalize_text(re.sub(r"[;,:-]+$", "", value[:cut_index]))
    if not truncated or len(truncated) < 24:
        return value
    return truncated


def _normalize_citation_text(text, prior_context_line=""):
    value = _normalize_text(text)
    if not value:
        return ""
    if _is_url_first_short_citation(value):
        prior = _normalize_text(prior_context_line)
        if prior and _looks_like_author_year_lead(prior):
            value = _normalize_text(f"{prior} {value}")

    value = _strip_double_slash_suffix(value)
    value = re.sub(r"\s+\/\/\s+", "; ", value)
    value = _trim_trailing_unmatched_closers(value)
    value = _repair_missing_trailing_closers(value)
    value = _truncate_citation_spillover(value)
    return value


def _normalize_citation_url_candidate(url_candidate):
    candidate = _normalize_text(url_candidate).rstrip(".,;:!?")
    if not candidate:
        return ""
    if candidate.lower().startswith("www."):
        candidate = f"https://{candidate}"
    parsed = urlparse(candidate)
    if parsed.scheme not in ("http", "https"):
        return ""
    if not parsed.netloc or "." not in parsed.netloc:
        return ""
    return candidate


def _extract_citation_url(citation):
    normalized = _normalize_text(citation)
    if not normalized:
        return ""
    for match in URL_CANDIDATE_RE.findall(normalized):
        candidate = _normalize_citation_url_candidate(match)
        if candidate:
            return candidate
    return ""


def _extract_prior_context_line(scored, citation_index):
    for item in reversed(scored):
        if item["index"] >= citation_index:
            continue
        if item["text"]:
            return item["text"]
    return ""


def _collect_body_content(
    body_indexes,
    paragraphs_by_index,
    include_underlined=True,
    include_highlighted=False,
):
    marked_runs = []
    underlined_runs = []
    highlighted_runs = []
    full_quote_lines = []

    for paragraph_index in body_indexes:
        para = paragraphs_by_index.get(paragraph_index)
        if para is None:
            continue
        para_text = (para.text or "").strip()
        if para_text:
            full_quote_lines.append(para_text)

        for run in para.runs:
            run_text = run.text or ""
            if not run_text.strip():
                continue
            is_underlined = bool(run.font.underline)
            highlight = run.font.highlight_color
            is_highlighted = bool(highlight) and highlight != WD_COLOR_INDEX.AUTO
            if (include_underlined and is_underlined) or (
                include_highlighted and is_highlighted
            ):
                marked_runs.append(run_text)
            if is_underlined:
                underlined_runs.append(run_text)
            if is_highlighted:
                highlighted_runs.append(run_text)

    return {
        "marked_text": " ".join(t.strip() for t in marked_runs if t.strip()),
        "underlined_text": " ".join(t.strip() for t in underlined_runs if t.strip()),
        "highlighted_text": " ".join(t.strip() for t in highlighted_runs if t.strip()),
        "full_quote": "\n".join(t.strip() for t in full_quote_lines if t.strip()),
    }


def parse_citation_metadata(citation_text):
    text = citation_text or ""
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    first_line = lines[0] if lines else ""

    oral = ""
    if first_line:
        if "[" in first_line and not first_line.lstrip().startswith("["):
            oral = first_line.split("[", 1)[0].strip()
        if not oral and first_line.lstrip().startswith("["):
            inner = first_line.lstrip("[").split("]", 1)[0]
            oral = inner.split(",", 1)[0].split(";", 1)[0].strip()
        if not oral and first_line.lstrip().startswith("("):
            inner = first_line.lstrip("(").split(")", 1)[0]
            oral = inner.split(",", 1)[0].split(";", 1)[0].strip()
        if not oral:
            dash_positions = []
            for dash in (" - ", " — ", " – ", "—", "–"):
                pos = first_line.find(dash)
                if pos != -1:
                    dash_positions.append(pos)
            dash_pos = min(dash_positions) if dash_positions else -1
            paren_pos = first_line.find("(")

            if paren_pos != -1 and (dash_pos == -1 or paren_pos < dash_pos):
                oral = first_line[:paren_pos].strip()
            elif dash_pos != -1:
                oral = first_line[:dash_pos].strip()
        if not oral:
            oral = first_line
        oral = oral.strip("[]() \t-—–")

    url = _extract_citation_url(text)

    return oral, url


def extract_cards(
    doc,
    title_level=None,
    tag_level=None,
    include_underlined=True,
    include_highlighted=False,
    parse_errors=None,
):
    del title_level, tag_level

    cards = []
    paragraphs = doc.paragraphs
    paragraphs_by_index = {}
    scored = []

    for i, para in enumerate(paragraphs):
        paragraphs_by_index[i] = para
        text = _normalize_text(para.text or "")
        if not text:
            continue
        try:
            scored.append(_build_paragraph_info(i, para))
        except Exception as e:
            if parse_errors is not None:
                snippet = text[:120]
                parse_errors.append(
                    f"Paragraph {i}: {e.__class__.__name__}: {e} | text={snippet!r}"
                )

    if not scored:
        return cards

    scored_by_index = {item["index"]: item for item in scored}
    require_url_for_citation = any(item["has_url"] for item in scored)
    citation_candidates = [
        item
        for item in scored
        if _is_likely_citation_paragraph(item, require_url_for_citation)
    ]

    if not citation_candidates:
        title = _reconstruct_title_from_body(scored)
        tag = _derive_tag_from_title(title, "") or title
        body_indexes = [item["index"] for item in scored]
        body_content = _collect_body_content(
            body_indexes,
            paragraphs_by_index,
            include_underlined=include_underlined,
            include_highlighted=include_highlighted,
        )
        if title or tag or body_content["marked_text"] or body_content["full_quote"]:
            cards.append(
                {
                    "title": title or "",
                    "tag": tag or "",
                    "citation": "",
                    "marked_text": body_content["marked_text"],
                    "underlined_text": body_content["underlined_text"],
                    "highlighted_text": body_content["highlighted_text"],
                    "full_quote": body_content["full_quote"],
                }
            )
        return cards

    title_selection_by_citation = {}
    citation_indexes = [item["index"] for item in citation_candidates]
    for idx, citation_index in enumerate(citation_indexes):
        previous_citation = citation_indexes[idx - 1] if idx > 0 else -1
        pre_range = [
            paragraph
            for paragraph in scored
            if previous_citation < paragraph["index"] < citation_index
        ]
        title_selection_by_citation[citation_index] = _select_title_from_pre_citation_range(
            pre_range, require_url_for_citation
        )

    for idx, citation_paragraph in enumerate(citation_candidates):
        citation_index = citation_paragraph["index"]
        try:
            next_citation = citation_indexes[idx + 1] if idx + 1 < len(citation_indexes) else None
            next_title_selection = (
                title_selection_by_citation.get(next_citation, {})
                if next_citation is not None
                else {}
            )
            next_title_indexes = next_title_selection.get("indexes", []) or []
            next_title_start = min(next_title_indexes) if next_title_indexes else None

            body_paragraphs = []
            for paragraph in scored:
                paragraph_index = paragraph["index"]
                if paragraph_index <= citation_index:
                    continue
                if next_title_start is not None and paragraph_index >= next_title_start:
                    continue
                if next_citation is not None and paragraph_index >= next_citation:
                    continue
                body_paragraphs.append(paragraph)

            title_selection = title_selection_by_citation.get(
                citation_index, {"title": "", "indexes": []}
            )
            selected_parts = _derive_selected_title_and_tag(title_selection, scored_by_index)
            prior_context_line = _extract_prior_context_line(scored, citation_index)
            citation_text = _normalize_citation_text(
                citation_paragraph["text"], prior_context_line=prior_context_line
            )
            citation_fallback = _derive_from_citation_lead(citation_text)

            title = selected_parts.get("title", "") or citation_fallback.get("title", "")
            if not title or _looks_like_tag_line(title):
                heading_title = _find_recent_heading_title(scored, citation_index)
                if heading_title:
                    title = heading_title
            if not title:
                title = _reconstruct_title_from_body(body_paragraphs)

            tag = selected_parts.get("tag", "")
            if not tag:
                tag = _derive_tag_from_title(title, citation_text)
            if not tag:
                tag = citation_fallback.get("tag", "") or title

            body_indexes = [item["index"] for item in body_paragraphs]
            body_content = _collect_body_content(
                body_indexes,
                paragraphs_by_index,
                include_underlined=include_underlined,
                include_highlighted=include_highlighted,
            )

            cards.append(
                {
                    "title": title or "",
                    "tag": tag or "",
                    "citation": citation_text or "",
                    "marked_text": body_content["marked_text"],
                    "underlined_text": body_content["underlined_text"],
                    "highlighted_text": body_content["highlighted_text"],
                    "full_quote": body_content["full_quote"],
                }
            )
        except Exception as e:
            if parse_errors is not None:
                snippet = citation_paragraph["text"][:120]
                parse_errors.append(
                    f"Citation paragraph {citation_index}: {e.__class__.__name__}: {e} | text={snippet!r}"
                )

    return cards


def format_txt(cards):
    lines = []
    for c in cards:
        if c["title"].strip():
            lines.append(c["title"].strip())
        if c["tag"].strip():
            lines.append(c["tag"].strip())
        if c["citation"].strip():
            lines.append(c["citation"].strip())
        if c["marked_text"].strip():
            lines.append(c["marked_text"].strip())
        lines.append("")
        lines.append("=" * 60)
        lines.append("")
    text = "\n".join(lines).rstrip() + "\n"
    return text


def format_csv(cards):
    output = io.StringIO()
    writer = csv.writer(output)
    for c in cards:
        col1 = c["title"].strip()
        parts = []
        if c["tag"].strip():
            parts.append(c["tag"].strip())
        if c["citation"].strip():
            parts.append(c["citation"].strip())
        if c["marked_text"].strip():
            parts.append(c["marked_text"].strip())
        col2 = "\n".join(parts)
        writer.writerow([col1, col2])
    return output.getvalue()


def format_csv_detailed(cards):
    raise NotImplementedError("Use format_csv_detailed_with_source instead.")


def format_csv_detailed_with_source(cards_with_source):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "filename",
            "school",
            "team_code",
            "side",
            "tournament_raw",
            "tournament",
            "round",
            "title",
            "tag",
            "citation",
            "oral_citation",
            "url",
            "highlighted",
            "underlined",
            "full_quote",
        ]
    )
    for source_name, c in cards_with_source:
        meta = parse_filename_metadata(source_name)
        oral, url = parse_citation_metadata(c["citation"])
        writer.writerow(
            [
                meta["filename"] or "",
                meta["school"] or "",
                meta["team_code"] or "",
                meta["side"] or "",
                meta["tournament_raw"] or "",
                meta["tournament"] or "",
                meta["round"] or "",
                c["title"].strip(),
                c["tag"].strip(),
                c["citation"].strip(),
                oral,
                url,
                c["highlighted_text"].strip(),
                c["underlined_text"].strip(),
                c["full_quote"].strip(),
            ]
        )
    return output.getvalue()


def _workbook_to_bytes(wb):
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def format_xlsx(cards):
    wb = Workbook()
    ws = wb.active
    ws.title = "cards"

    ws.append(["title", "card"])
    for c in cards:
        parts = []
        if c["tag"].strip():
            parts.append(c["tag"].strip())
        if c["citation"].strip():
            parts.append(c["citation"].strip())
        if c["marked_text"].strip():
            parts.append(c["marked_text"].strip())
        card_text = "\n".join(parts)
        ws.append([c["title"].strip(), card_text])

    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = wrap

    return _workbook_to_bytes(wb)


def format_xlsx_detailed_with_source(cards_with_source):
    wb = Workbook()
    ws = wb.active
    ws.title = "cards_detailed"

    ws.append(
        [
            "filename",
            "school",
            "team_code",
            "side",
            "tournament_raw",
            "tournament",
            "round",
            "title",
            "tag",
            "citation",
            "oral_citation",
            "url",
            "highlighted",
            "underlined",
            "full_quote",
        ]
    )
    for source_name, c in cards_with_source:
        meta = parse_filename_metadata(source_name)
        oral, url = parse_citation_metadata(c["citation"])
        ws.append(
            [
                meta["filename"] or "",
                meta["school"] or "",
                meta["team_code"] or "",
                meta["side"] or "",
                meta["tournament_raw"] or "",
                meta["tournament"] or "",
                meta["round"] or "",
                c["title"].strip(),
                c["tag"].strip(),
                c["citation"].strip(),
                oral,
                url,
                c["highlighted_text"].strip(),
                c["underlined_text"].strip(),
                c["full_quote"].strip(),
            ]
        )

    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=15):
        for cell in row:
            cell.alignment = wrap

    return _workbook_to_bytes(wb)
